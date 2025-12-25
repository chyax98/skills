#!/usr/bin/env python3
"""
将 JSONL 测试用例转换为 Excel 格式

用法：
    python3 to_excel.py input.jsonl -o output.xlsx
"""

import json
import sys
import argparse
from pathlib import Path

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("错误：需要安装 openpyxl 库")
    print("执行：pip install openpyxl")
    sys.exit(1)


def read_jsonl(file_path):
    """读取 JSONL 文件"""
    cases = []
    with open(file_path, 'r', encoding='utf-8') as f:
        for line in f:
            line = line.strip()
            if line:
                cases.append(json.loads(line))
    return cases


def format_steps_actions(steps):
    """格式化操作步骤（只包含 action）"""
    if not steps:
        return ""

    formatted = []
    for i, step in enumerate(steps, start=1):
        action = step.get('action', '')
        formatted.append(f"{i}. {action}")

    return "\n".join(formatted)


def format_steps_expected(steps):
    """格式化预期结果（只包含 expected）"""
    if not steps:
        return ""

    formatted = []
    step_num = 1
    for step in steps:
        if 'expected' in step:
            expected = step.get('expected', '')
            formatted.append(f"{step_num}. {expected}")
            step_num += 1

    return "\n".join(formatted)


def to_excel(input_file, output_file):
    """将 JSONL 转换为 Excel"""

    # 读取用例
    cases = read_jsonl(input_file)

    if not cases:
        print("❌ 没有找到任何用例")
        return

    # 创建工作簿
    wb = Workbook()
    ws = wb.active
    ws.title = "测试用例"

    # 定义列头（按用户要求）
    headers = [
        "一级模块",      # module_name
        "编号",          # 自动生成序号
        "测试项",        # test_item
        "优先级",        # priority
        "用例标题",      # name
        "操作步骤",      # steps (action only)
        "预期结果",      # steps (expected only)
        "是否反向用例",  # is_negative (是/否)
        "测试类型",      # test_type
        "AI生成",        # 固定值"是"
        "备注"           # notes
    ]

    # 写入表头
    ws.append(headers)

    # 设置表头样式
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment

    # 设置边框样式
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # 按模块分组
    cases_by_module = {}
    for case in cases:
        module_name = case.get("module_name", "未分类")
        if module_name not in cases_by_module:
            cases_by_module[module_name] = []
        cases_by_module[module_name].append(case)

    # 全局编号
    global_index = 1

    # 写入数据（按模块分组）
    for module_name in sorted(cases_by_module.keys()):
        module_cases = cases_by_module[module_name]

        for case in module_cases:
            # 提取字段
            test_item = case.get("test_item", "")
            priority = case.get("priority", "")
            name = case.get("name", "")
            steps = case.get("steps", [])
            is_negative = case.get("is_negative", False)
            test_type = case.get("test_type", "")
            notes = case.get("notes", "")

            # 格式化操作步骤和预期结果
            formatted_steps = format_steps_actions(steps)
            formatted_expected = format_steps_expected(steps)

            # 转换是否反向用例
            is_negative_text = "是" if is_negative else "否"

            # 写入行
            row_data = [
                module_name,         # 一级模块
                global_index,        # 编号
                test_item,           # 测试项
                priority,            # 优先级
                name,                # 用例标题
                formatted_steps,     # 操作步骤
                formatted_expected,  # 预期结果
                is_negative_text,    # 是否反向用例
                test_type,           # 测试类型
                "是",                # AI生成（固定值）
                notes                # 备注
            ]

            ws.append(row_data)
            global_index += 1

    # 设置列宽
    column_widths = {
        'A': 15,   # 一级模块
        'B': 8,    # 编号
        'C': 20,   # 测试项
        'D': 10,   # 优先级
        'E': 50,   # 用例标题
        'F': 60,   # 操作步骤
        'G': 60,   # 预期结果
        'H': 12,   # 是否反向用例
        'I': 15,   # 测试类型
        'J': 10,   # AI生成
        'K': 30    # 备注
    }

    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width

    # 设置数据行样式
    for row in range(2, ws.max_row + 1):
        for col in range(1, len(headers) + 1):
            cell = ws.cell(row=row, column=col)

            # 边框
            cell.border = thin_border

            # 对齐方式
            if col in [1, 2, 3, 4, 8, 9, 10]:  # 模块、编号、测试项、优先级、是否反向、测试类型、AI生成
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            else:  # 用例标题、操作步骤、预期结果、备注
                cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)

    # 冻结首行
    ws.freeze_panes = "A2"

    # 添加筛选器
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{ws.max_row}"

    # 保存文件
    wb.save(output_file)

    print(f"✅ Excel 导出完成：{len(cases)} 条用例")
    print(f"   文件路径：{output_file}")
    print(f"   模块数量：{len(cases_by_module)}")


def main():
    parser = argparse.ArgumentParser(description="将 JSONL 测试用例转换为 Excel 格式")
    parser.add_argument("input", type=Path, help="输入 JSONL 文件")
    parser.add_argument("-o", "--output", type=Path, required=True, help="输出 Excel 文件")

    args = parser.parse_args()

    # 检查输入文件
    if not args.input.exists():
        print(f"❌ 文件不存在：{args.input}")
        sys.exit(1)

    try:
        to_excel(args.input, args.output)
    except Exception as e:
        print(f"❌ 转换失败：{e}")
        import traceback
        traceback.print_exc()
        sys.exit(1)


if __name__ == "__main__":
    main()
