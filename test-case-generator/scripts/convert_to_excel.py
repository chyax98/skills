
"""
测试用例 Markdown 转 Excel 工具

将 Markdown 格式的测试用例转换为 Excel 格式。

使用方法：
    uv run convert_to_excel.py test-cases.md -o test-cases.xlsx
    uv run convert_to_excel.py test-cases.md  # 默认输出同名 xlsx 文件
"""

import re
import argparse
from pathlib import Path
from dataclasses import dataclass, field
from typing import Optional


@dataclass
class TestCase:
    """测试用例数据结构"""
    module: str = ""              # 一级分组
    scenario: str = ""            # 二级分组
    name: str = ""                # 用例名称
    priority: str = ""            # 优先级
    test_type: str = ""           # 测试类型
    is_negative: bool = False     # 是否反向
    preconditions: str = ""       # 前置条件
    steps: list = field(default_factory=list)  # [(action, expected), ...]
    notes: str = ""               # 备注


def parse_metadata(line: str) -> tuple[str, str, bool]:
    """
    解析元数据行：P1|功能|反向
    返回：(priority, test_type, is_negative)
    """
    parts = [p.strip() for p in line.split('|')]
    priority = parts[0] if parts else ""
    test_type = parts[1] if len(parts) > 1 else ""
    is_negative = len(parts) > 2 and '反向' in parts[2]
    return priority, test_type, is_negative


def parse_steps(lines: list[str]) -> list[tuple[str, str]]:
    """
    解析步骤行
    格式：1. 操作 ➡️ 预期结果  或  1. 操作
    返回：[(action, expected), ...]
    """
    steps = []
    for line in lines:
        line = line.strip()
        if not line:
            continue

        # 移除编号前缀
        match = re.match(r'^\d+\.\s*(.+)$', line)
        if match:
            content = match.group(1)

            # 使用 ➡️ 作为分隔符
            if '➡️' in content:
                parts = content.split('➡️', 1)
                action = parts[0].strip()
                expected = parts[1].strip() if len(parts) > 1 else ""
            else:
                action = content.strip()
                expected = ""

            steps.append((action, expected))

    return steps


def parse_markdown(content: str) -> list[TestCase]:
    """解析 Markdown 内容，返回测试用例列表"""
    cases = []
    lines = content.split('\n')

    current_module = ""
    current_scenario = ""
    current_case: Optional[TestCase] = None
    current_section = None  # 当前解析的部分：preconditions, steps, notes
    section_lines = []

    def save_section():
        """保存当前部分到用例"""
        nonlocal current_case, current_section, section_lines
        if current_case and current_section and section_lines:
            if current_section == 'preconditions':
                # 前置条件：合并为单行，用分号分隔
                precond_list = []
                for line in section_lines:
                    line = line.strip()
                    if line.startswith('- '):
                        line = line[2:]
                    if line:
                        precond_list.append(line)
                current_case.preconditions = '；'.join(precond_list)
            elif current_section == 'steps':
                current_case.steps = parse_steps(section_lines)
            elif current_section == 'notes':
                current_case.notes = ' '.join(line.strip() for line in section_lines if line.strip())
        section_lines = []
        current_section = None

    def save_case():
        """保存当前用例"""
        nonlocal current_case
        save_section()
        if current_case and current_case.name:
            cases.append(current_case)
        current_case = None

    for line in lines:
        stripped = line.strip()

        # 跳过分隔线
        if stripped == '---':
            continue

        # 一级标题：模块
        if line.startswith('# ') and not line.startswith('## '):
            save_case()
            current_module = line[2:].strip()
            current_scenario = ""
            continue

        # 二级标题：场景
        if line.startswith('## '):
            save_case()
            current_scenario = line[3:].strip()
            continue

        # 三级标题：用例名称
        if line.startswith('### '):
            save_case()
            current_case = TestCase(
                module=current_module,
                scenario=current_scenario,
                name=line[4:].strip()
            )
            continue

        # 元数据：P1|功能|反向
        if current_case and re.match(r'^P[1-5]\|', stripped):
            save_section()
            priority, test_type, is_negative = parse_metadata(stripped)
            current_case.priority = priority
            current_case.test_type = test_type
            current_case.is_negative = is_negative
            continue

        # 前置条件
        if stripped.startswith('[前置]'):
            if current_case:
                save_section()
                current_section = 'preconditions'
                # 如果同一行有内容
                rest = stripped[4:].strip()
                if rest:
                    section_lines.append(rest)
            continue

        # 步骤
        if stripped.startswith('[步骤]'):
            if current_case:
                save_section()
                current_section = 'steps'
            continue

        # 备注
        if stripped.startswith('[备注]'):
            if current_case:
                save_section()
                current_section = 'notes'
                rest = stripped[4:].strip()
                if rest:
                    section_lines.append(rest)
            continue

        # 收集当前部分的内容
        if current_section and stripped:
            section_lines.append(stripped)

    # 保存最后一个用例
    save_case()

    return cases


def create_excel(cases: list[TestCase], output_path: Path):
    """创建 Excel 文件"""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    except ImportError:
        print("错误：需要安装 openpyxl")
        print("请执行：pip install openpyxl")
        return False

    wb = Workbook()
    ws = wb.active
    ws.title = "测试用例"

    # 定义样式
    header_font = Font(bold=True, size=11)
    header_fill = PatternFill(start_color="CCE5FF", end_color="CCE5FF", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell_alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # 表头
    headers = [
        "一级分组", "二级分组", "用例名称", "优先级", "测试类型",
        "是否反向", "前置条件", "操作步骤", "预期结果", "备注"
    ]

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    # 数据行
    for row_idx, case in enumerate(cases, 2):
        # 步骤和预期结果格式化
        steps_text = '\n'.join(f"{i+1}. {step[0]}" for i, step in enumerate(case.steps))
        expected_text = '\n'.join(f"{i+1}. {step[1]}" for i, step in enumerate(case.steps) if step[1])

        row_data = [
            case.module,
            case.scenario,
            case.name,
            case.priority,
            case.test_type,
            "是" if case.is_negative else "否",
            case.preconditions,
            steps_text,
            expected_text,
            case.notes
        ]

        for col, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col, value=value)
            cell.alignment = cell_alignment
            cell.border = thin_border

    # 设置列宽
    column_widths = [15, 15, 30, 8, 12, 10, 25, 35, 35, 20]
    for col, width in enumerate(column_widths, 1):
        ws.column_dimensions[chr(64 + col)].width = width

    # 冻结首行
    ws.freeze_panes = 'A2'

    # 保存
    wb.save(output_path)
    return True


def main():
    parser = argparse.ArgumentParser(
        description='将 Markdown 格式的测试用例转换为 Excel',
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog='''
示例：
    python convert_to_excel.py test-cases.md -o test-cases.xlsx
    python convert_to_excel.py test-cases.md  # 输出 test-cases.xlsx
        '''
    )
    parser.add_argument('input', help='输入的 Markdown 文件路径')
    parser.add_argument('-o', '--output', help='输出的 Excel 文件路径（默认同名 .xlsx）')

    args = parser.parse_args()

    input_path = Path(args.input)
    if not input_path.exists():
        print(f"错误：文件不存在 {input_path}")
        return 1

    output_path = Path(args.output) if args.output else input_path.with_suffix('.xlsx')

    # 读取并解析
    content = input_path.read_text(encoding='utf-8')
    cases = parse_markdown(content)

    if not cases:
        print("警告：未解析到任何测试用例")
        return 1

    # 创建 Excel
    if create_excel(cases, output_path):
        print(f"转换完成：{output_path}")
        print(f"共 {len(cases)} 条用例")
        return 0
    else:
        return 1


if __name__ == '__main__':
    exit(main())
