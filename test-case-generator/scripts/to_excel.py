#!/usr/bin/env python3
"""
JSONL 转 Excel 工具

功能：
1. 将 JSONL 格式的测试用例转换为 Excel 格式
2. 自动调整列宽
3. 添加筛选器

用法：
    python convert_to_excel.py cases.jsonl -o cases.xlsx

依赖：
    pip install openpyxl
"""

import json
import sys
import argparse
from pathlib import Path
from typing import List, Dict

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("错误：需要安装 openpyxl 库")
    print("执行：pip install openpyxl")
    sys.exit(1)


# 列定义
COLUMNS = [
    ('id', '用例ID', 12),
    ('name', '用例名称', 50),
    ('module_name', '模块', 15),
    ('test_item', '测试项', 20),
    ('priority', '优先级', 10),
    ('test_type', '测试类型', 15),
    ('is_negative', '反向用例', 10),
    ('preconditions', '前置条件', 40),
    ('steps', '测试步骤', 60),
    ('notes', '备注', 30),
]

# 优先级颜色
PRIORITY_COLORS = {
    'P1': 'FF6B6B',  # 红色
    'P2': 'FFB347',  # 橙色
    'P3': 'FFD93D',  # 黄色
    'P4': '6BCB77',  # 绿色
    'P5': 'A8A8A8',  # 灰色
}


def load_jsonl(file_path: Path) -> List[Dict]:
    """加载 JSONL 文件"""
    objects = []
    try:
        with open(file_path, 'r', encoding='utf-8') as f:
            for line_num, line in enumerate(f, 1):
                line = line.strip()
                if not line:
                    continue

                try:
                    obj = json.loads(line)
                    objects.append(obj)
                except json.JSONDecodeError as e:
                    print(f"警告：{file_path}:{line_num} JSON 解析失败：{e}", file=sys.stderr)
                    continue
    except Exception as e:
        print(f"错误：无法读取文件 {file_path}：{e}", file=sys.stderr)
        sys.exit(1)

    return objects


def format_preconditions(preconditions: List[str]) -> str:
    """格式化前置条件"""
    if not preconditions:
        return ""
    return "\n".join(f"{i+1}. {p}" for i, p in enumerate(preconditions))


def format_steps(steps: List[Dict]) -> str:
    """格式化测试步骤"""
    if not steps:
        return ""
    lines = []
    for i, step in enumerate(steps, 1):
        action = step.get('action', '')
        expected = step.get('expected', '')
        lines.append(f"{i}. {action}")
        if expected:
            lines.append(f"   预期：{expected}")
    return "\n".join(lines)


def convert_to_excel(test_cases: List[Dict], output_path: Path):
    """转换测试用例为 Excel"""
    wb = Workbook()
    ws = wb.active
    ws.title = "测试用例"

    # 表头样式
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")

    # 写入表头
    for col_idx, (field, title, width) in enumerate(COLUMNS, 1):
        cell = ws.cell(row=1, column=col_idx, value=title)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # 边框样式
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # 写入数据
    for row_idx, case in enumerate(test_cases, 2):
        for col_idx, (field, _, _) in enumerate(COLUMNS, 1):
            value = case.get(field, '')

            # 特殊字段处理
            if field == 'preconditions':
                value = format_preconditions(value) if isinstance(value, list) else str(value)
            elif field == 'steps':
                value = format_steps(value) if isinstance(value, list) else str(value)
            elif field == 'is_negative':
                value = '是' if value else '否'

            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = thin_border
            cell.alignment = Alignment(vertical="top", wrap_text=True)

            # 优先级颜色
            if field == 'priority' and value in PRIORITY_COLORS:
                cell.fill = PatternFill(
                    start_color=PRIORITY_COLORS[value],
                    end_color=PRIORITY_COLORS[value],
                    fill_type="solid"
                )

    # 冻结首行
    ws.freeze_panes = 'A2'

    # 添加筛选器
    ws.auto_filter.ref = ws.dimensions

    # 保存
    wb.save(output_path)


def main():
    parser = argparse.ArgumentParser(
        description='JSONL 转 Excel 工具',
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
示例：
  python convert_to_excel.py cases.jsonl -o cases.xlsx
        """
    )
    parser.add_argument('input', type=Path, help='输入 JSONL 文件')
    parser.add_argument('-o', '--output', type=Path, required=True, help='输出 Excel 文件路径')

    args = parser.parse_args()

    # 检查输入文件
    if not args.input.exists():
        print(f"错误：文件不存在 {args.input}")
        sys.exit(1)

    # 加载数据
    test_cases = load_jsonl(args.input)

    if not test_cases:
        print("错误：JSONL 文件为空")
        sys.exit(1)

    # 确保输出目录存在
    args.output.parent.mkdir(parents=True, exist_ok=True)

    # 转换
    convert_to_excel(test_cases, args.output)

    print(f"✅ 已转换 {len(test_cases)} 条用例到 {args.output}")


if __name__ == '__main__':
    main()
